#! python
# EmergencyLabelFinder.py - separates cells searched for from those that don't meet criteria

import openpyxl, pprint, time
excelSheetName = input('Enter Name of Excel Sheet: ')
startingLetters = []
startingLetter = ''
print("Enter starting letters to filter onto a seperate tab. Enter 'done' when finished.")

while startingLetter != 'done':
    startingLetter = input("Enter a starting letter: ")
    if startingLetter != 'done':
        startingLetter = startingLetter.upper()
        startingLetters.append(startingLetter)

wb = openpyxl.load_workbook(excelSheetName + '.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
createdSheet = wb.create_sheet()
createdSheet.title = 'Labels'
createdEMsheet = wb.create_sheet()
createdEMsheet.title = 'EM Labels'
next_EM_row = 1
next_row = 1

for row in range(1, sheet.max_row + 1):
    label = sheet['A' + str(row)].value
    labelString = str(label) # This variable might not be needed
    if labelString.startswith(tuple(startingLetters)):
        createdEMsheet.cell(column = 1, row = next_EM_row, value = label)
        next_EM_row += 1
    elif label: # Using implicit booleanness of label object. If false skip
        createdSheet.cell(column = 1, row = next_row, value = label)
        next_row += 1

wb.save(excelSheetName + '.xlsx')
print('------------------------Done------------------------')
time.sleep(3)
