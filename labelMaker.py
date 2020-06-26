# first column is # needed, second column is label text

import openpyxl, pprint
excelSheetName = raw_input('Enter Name of Excel Sheet: ')

wb = openpyxl.load_workbook(excelSheetName + '.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
createdSheet = wb.create_sheet()
createdSheet.title = 'Labels'
next_row = 1

for row in range(1, sheet.max_row + 1):
    numberToMake = sheet['A' + str(row)].value
    labelContents = sheet['B' + str(row)].value
    i = 0
    while i < numberToMake:
        createdSheet.cell(column=1,row=next_row,value=labelContents)
        next_row += 1
        i += 1
print('------------------------Done------------------------')
wb.save(excelSheetName + '.xlsx')
