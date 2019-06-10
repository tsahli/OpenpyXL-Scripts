#!python
# Orders labels from Box Material Calulator Excel in aphabetical order

import openpyxl
from openpyxl import Workbook

print("This program will create a seperate Excel file containing only labels, sorted alphabetically.\nYou will find the label file in the same directory as the file you're creating labels from.\nThis program assumes all the labels you need start with a W.\n")

while True:
    try:
        excelSheetName = raw_input("Enter name of Excel Sheet: ")
        wb = openpyxl.load_workbook(excelSheetName + ".xlsx", data_only=True, read_only=True)
        break
    except:
        print("\n--------------------------------------\nWorkbook or sheet cannot be found!\nDouble check workbook spelling.\n--------------------------------------\n")

sheet = wb.get_sheet_by_name("LABELS")
print("Workbook and sheet loaded. Creating labels. This process can take some time.")

labelList = []

for row in range(2, sheet.max_row + 1):
    label = sheet["B" + str(row)].value
    labelString = str(label)
    if labelString.startswith("W"): # Assumes all needed labels start with "W"
        labelList.append(labelString)

print("Sorting labels and writing to new file.")
labelList.sort()
labelWorkbook = Workbook()
labelWorksheet = labelWorkbook.active
labelWorksheet.title = "Labels"
labelWorksheet.cell(row = 1, column = 1).value = "PRINT ME"
next_row = 2

for x in labelList:
    labelWorksheet.cell(column = 1, row = next_row, value = x)
    next_row += 1

labelWorkbook.save(excelSheetName + " LABELS" + ".xlsx")
print('------------------------Done------------------------')
print("Saved as: " + excelSheetName + " LABELS" + ".xlsx")
raw_input("Press ENTER to exit")


